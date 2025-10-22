"""
Excel报表生成模块

提供图表、样式和模板化报表生成功能
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)


class ExcelReporter:
    """Excel报表生成器"""

    def __init__(self):
        """初始化报表生成器"""
        self.wb = None
        self.ws = None

    def create_report(
        self,
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
        output: Union[str, Path],
        config: Optional[Dict[str, Any]] = None
    ) -> None:
        """创建完整报表

        Args:
            data: 数据，可以是DataFrame或{sheet_name: DataFrame}字典
            output: 输出路径
            config: 报表配置
                - title: 报表标题
                - sheets: 工作表配置列表
                - charts: 图表配置列表
                - styles: 样式配置

        Example:
            >>> reporter = ExcelReporter()
            >>> reporter.create_report(
            ...     data={'销售': sales_df, '成本': cost_df},
            ...     output='report.xlsx',
            ...     config={
            ...         'title': '月度经营报表',
            ...         'charts': [{'type': 'bar', 'sheet': '销售', 'x': '门店', 'y': '销售额'}]
            ...     }
            ... )
        """
        config = config or {}
        output = Path(output)
        output.parent.mkdir(parents=True, exist_ok=True)

        # 准备数据
        if isinstance(data, pd.DataFrame):
            sheets_data = {'Sheet1': data}
        else:
            sheets_data = data

        # 写入基础数据
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 加载工作簿进行格式化
        self.wb = load_workbook(output)

        # 应用样式
        if config.get('styles'):
            self._apply_styles(config['styles'])

        # 添加图表
        if config.get('charts'):
            self._add_charts(config['charts'], sheets_data)

        # 添加标题
        if config.get('title'):
            self._add_title(config['title'])

        # 保存
        self.wb.save(output)
        logger.info(f"报表创建完成: {output}")

    def add_charts(
        self,
        file_path: Union[str, Path],
        data: pd.DataFrame,
        charts: List[Dict[str, Any]],
        sheet_name: str = 'Sheet1'
    ) -> None:
        """为现有Excel文件添加图表

        Args:
            file_path: Excel文件路径
            data: 数据
            charts: 图表配置列表
            sheet_name: 工作表名称

        Example:
            >>> reporter.add_charts(
            ...     'report.xlsx',
            ...     data,
            ...     charts=[
            ...         {'type': 'bar', 'x': '门店', 'y': '销售额', 'position': 'E2'},
            ...         {'type': 'line', 'x': '月份', 'y': '销售额', 'position': 'E15'}
            ...     ]
            ... )
        """
        self.wb = load_workbook(file_path)
        self.ws = self.wb[sheet_name]

        for chart_config in charts:
            self._create_chart(chart_config, data)

        self.wb.save(file_path)
        logger.info(f"图表添加完成: {len(charts)} 个图表")

    def _create_chart(self, config: Dict[str, Any], data: pd.DataFrame) -> None:
        """创建单个图表

        Args:
            config: 图表配置
                - type: 图表类型 ('bar', 'line', 'pie')
                - x: X轴列名
                - y: Y轴列名
                - position: 图表位置（如 'E2'）
                - title: 图表标题
            data: 数据
        """
        chart_type = config.get('type', 'bar')
        position = config.get('position', 'E2')

        # 获取数据范围
        max_row = self.ws.max_row
        x_col = self._get_column_index(config.get('x'))
        y_col = self._get_column_index(config.get('y'))

        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            logger.warning(f"不支持的图表类型: {chart_type}")
            return

        # 设置数据
        if chart_type in ['bar', 'line']:
            # 类别轴（X轴）
            cats = Reference(self.ws, min_col=x_col, min_row=2, max_row=max_row)
            # 数值轴（Y轴）
            values = Reference(self.ws, min_col=y_col, min_row=1, max_row=max_row)

            chart.add_data(values, titles_from_data=True)
            chart.set_categories(cats)
        else:  # pie
            values = Reference(self.ws, min_col=y_col, min_row=2, max_row=max_row)
            labels = Reference(self.ws, min_col=x_col, min_row=2, max_row=max_row)

            chart.add_data(values)
            chart.set_categories(labels)

        # 设置标题
        if 'title' in config:
            chart.title = config['title']

        # 添加到工作表
        self.ws.add_chart(chart, position)

    def _apply_styles(self, styles_config: Dict[str, Any]) -> None:
        """应用样式配置

        Args:
            styles_config: 样式配置
                - header: 表头样式
                - data: 数据样式
                - title: 标题样式
        """
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            # 应用表头样式
            if 'header' in styles_config:
                header_style = styles_config['header']
                for cell in ws[1]:
                    self._apply_cell_style(cell, header_style)

            # 应用数据样式
            if 'data' in styles_config:
                data_style = styles_config['data']
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        self._apply_cell_style(cell, data_style)

    def _apply_cell_style(self, cell, style_config: Dict[str, Any]) -> None:
        """应用单元格样式

        Args:
            cell: openpyxl单元格对象
            style_config: 样式配置
                - font: 字体配置 {'name': '微软雅黑', 'size': 12, 'bold': True, 'color': 'FFFFFF'}
                - fill: 填充配置 {'color': '4472C4'}
                - alignment: 对齐配置 {'horizontal': 'center', 'vertical': 'center'}
                - border: 边框配置 {'style': 'thin', 'color': '000000'}
        """
        # 字体
        if 'font' in style_config:
            font_cfg = style_config['font']
            cell.font = Font(
                name=font_cfg.get('name', '微软雅黑'),
                size=font_cfg.get('size', 11),
                bold=font_cfg.get('bold', False),
                color=font_cfg.get('color')
            )

        # 填充
        if 'fill' in style_config:
            fill_cfg = style_config['fill']
            cell.fill = PatternFill(
                start_color=fill_cfg.get('color', 'FFFFFF'),
                end_color=fill_cfg.get('color', 'FFFFFF'),
                fill_type='solid'
            )

        # 对齐
        if 'alignment' in style_config:
            align_cfg = style_config['alignment']
            cell.alignment = Alignment(
                horizontal=align_cfg.get('horizontal', 'left'),
                vertical=align_cfg.get('vertical', 'center')
            )

        # 边框
        if 'border' in style_config:
            border_cfg = style_config['border']
            side = Side(
                style=border_cfg.get('style', 'thin'),
                color=border_cfg.get('color', '000000')
            )
            cell.border = Border(left=side, right=side, top=side, bottom=side)

    def _add_title(self, title: str) -> None:
        """添加报表标题

        Args:
            title: 标题文本
        """
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            # 在第一行插入标题
            ws.insert_rows(1)
            ws['A1'] = title

            # 合并单元格
            ws.merge_cells(f'A1:{self._get_column_letter(ws.max_column)}1')

            # 标题样式
            title_cell = ws['A1']
            title_cell.font = Font(name='微软雅黑', size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

    def _get_column_index(self, column_name: str) -> int:
        """根据列名获取列索引

        Args:
            column_name: 列名

        Returns:
            int: 列索引（从1开始）
        """
        if not self.ws:
            return 1

        for idx, cell in enumerate(self.ws[1], start=1):
            if cell.value == column_name:
                return idx
        return 1

    def _get_column_letter(self, col_idx: int) -> str:
        """根据列索引获取列字母

        Args:
            col_idx: 列索引（从1开始）

        Returns:
            str: 列字母（如 'A', 'B', 'AA'）
        """
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_idx)


# 预定义样式
DEFAULT_STYLES = {
    'header': {
        'font': {'name': '微软雅黑', 'size': 12, 'bold': True, 'color': 'FFFFFF'},
        'fill': {'color': '4472C4'},
        'alignment': {'horizontal': 'center', 'vertical': 'center'},
        'border': {'style': 'thin', 'color': '000000'}
    },
    'data': {
        'font': {'name': '微软雅黑', 'size': 11},
        'alignment': {'horizontal': 'left', 'vertical': 'center'},
        'border': {'style': 'thin', 'color': 'D0D0D0'}
    }
}
